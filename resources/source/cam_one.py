from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from pymediainfo import MediaInfo
from openpyxl import Workbook, load_workbook
import requests
import time
import logging
import os
import sys
import io

# 로봇용 전역
driver = None
def init_driver():
    global driver
    driver = setup_driver()

# 터미널에 로그출력
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 메인페이지
def setup_driver():
    options = Options()
    options.add_experimental_option("detach", True)
    # 이게 더 안정적이긴 함 ㅎㅎ
    # options.add_argument("--headless")
    # 과시용 
    options.add_argument("--window-size=1920,1080")
    service = Service()
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(3)
    return driver

# 로그인
def login(ip, username, password):
    driver.get(ip)
    action = ActionChains(driver)
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[1]/div[2]/div/form/p[1]/input').click()
    action.send_keys(username).key_down(Keys.TAB).send_keys(password).pause(1).key_down(Keys.ENTER).perform()
    action.reset_actions()
    time.sleep(5)
    logger.info('login success')

# 설정로그인 + 비디오 스트림 3, 4 h.264로 설정
def loginop(ip, username, password):
    driver.get(ip)
    time.sleep(5)
    try:
        element = driver.find_element(By.CSS_SELECTOR, '#user')
        element.click()
        element.send_keys(username)
        time.sleep(3)
        element = driver.find_element(By.CSS_SELECTOR, '#pwd')
        element.click()
        element.send_keys(password)
        time.sleep(3)
        element = driver.find_element(By.CSS_SELECTOR, '#login_form > p:nth-child(8) > div > ins')
        element.click()
        element = driver.find_element(By.CSS_SELECTOR, '#login_submit')
        element.click()
    except Exception as e:
        print(f"Error: {e}")
    time.sleep(5)
    logger.info('option menu login complete')
    time.sleep(5)
    try:
        element = driver.find_element(By.CSS_SELECTOR, '#sidebar > ul > li:nth-child(2)')
        element.click()
        time.sleep(3)
        element = driver.find_element(By.CSS_SELECTOR, '#sidebar > ul > li.open > ul > li:nth-child(2) > a')
        element.click()
        time.sleep(3)
        select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_codec_list'))
        select.select_by_value("H264HP")
        time.sleep(3)
        select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_codec_list'))
        select.select_by_value("H264HP")
        time.sleep(3)
        element = driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]')
        element.click()
    except Exception as e:
        print(f"Error: {e}")
    logger.info('Stream 3, 4 Codec Change Complete')

# 탭 호출 > 라이브
def tab_l():
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[1]/div/button').click()
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[2]/div[1]/div[1]/div[1]/a[1]').click()
    logger.info('Go to Live Mode')

# 탭 호출 > 플레이백
def tab_p():
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[1]/div/button').click()
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[2]/div[1]/div[1]/div[1]/a[2]').click()
    logger.info('Go to Playback Mode')

# 탭 호출 > 설정
def tab_o():
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[1]/div/button').click()
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[2]/div[1]/div[1]/div[1]/a[3]').click()
    logger.info('Go to Options')

# 탭 호출 > 로그아웃
def tab_bye():
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[1]/div/button').click()
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[2]/div[1]/div[1]/div[1]/a[4]').click()
    logger.info('Logout Complete')

# 설정 > 비디오 스트림 진입
def op_stream():
    time.sleep(5)
    element = driver.find_element(By.CSS_SELECTOR, '#sidebar > ul > li:nth-child(2)')
    element.click()
    time.sleep(3)
    element = driver.find_element(By.CSS_SELECTOR, '#sidebar > ul > li.open > ul > li:nth-child(2) > a')
    element.click()
    time.sleep(3)
    logger.info('Options > Video Stream Now')

# 플레이백 > 비디오 스트림 진입
def op_stream_pb():
    try:
        element = driver.find_element(By.XPATH, '//*[@id="menu_button"]')
        element.click()
        element = driver.find_element(By.XPATH, '//*[@id="sidebar-shortcuts-large"]/a[3]')
        element.click()
        time.sleep(3)
        element = driver.find_element(By.CSS_SELECTOR, '#sidebar > ul > li:nth-child(2)')
        element.click()
        time.sleep(3)
        element = driver.find_element(By.CSS_SELECTOR, '#sidebar > ul > li.open > ul > li:nth-child(2) > a')
        element.click()
        time.sleep(3)
    except Exception as e:
        print(f"Error: {e}")
    logger.info('Playback > Video Stream Now')


# 저장소 겹쳐쓰기 설정
def reckeep():
    element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[4]/a')
    element.click()
    element = driver.find_element(By.XPATH, '//*[@id="tab-record"]/div[1]/label/div')
    element.click()
    element = driver.find_element(By.XPATH, '//*[@id="main-container"]/div[2]/div/div[2]/div[1]/h1/button')
    element.click()
    logger.info('Keep Save Mode On')

# 이벤트 녹화 사용설정-비디오 스트림1
def recon():
    try:
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/a')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/ul/li[2]/a')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="tab-record"]/div[1]/label/div/ins')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="main-container"]/div[2]/div/div[2]/div[1]/h1/button')
        element.click()
        time.sleep(5)
    except Exception as e:
        print(f"Error:{e}")
    logger.info('Event Rec On with Video Stream1')

# 이벤트 녹화 사용설정-비디오 스트림2
def recon2():
    try:
        element = driver.find_element(By.XPATH, '//*[@id="menu_button"]')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar-shortcuts-large"]/a[3]')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/a')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/ul/li[2]/a')
        element.click()
        time.sleep(5)
        select = Select(driver.find_element(By.CSS_SELECTOR,'#real_record_event_video'))
        select.select_by_value('2')
        element = driver.find_element(By.XPATH, '//*[@id="main-container"]/div[2]/div/div[2]/div[1]/h1/button')
        element.click()
        time.sleep(5)
    except Exception as e:
        print(f"Error:{e}")
    logger.info('Event Rec On with Video Stream2')

# 이벤트 녹화 사용설정-비디오 스트림3
def recon3():
    try:
        # element = driver.find_element(By.XPATH, '//*[@id="menu_button"]')
        # element.click()
        # time.sleep(5)
        # element = driver.find_element(By.XPATH, '//*[@id="sidebar-shortcuts-large"]/a[3]')
        # element.click()
        # time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/a')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/ul/li[2]/a')
        element.click()
        time.sleep(5)
        select = Select(driver.find_element(By.CSS_SELECTOR,'#real_record_event_video'))
        select.select_by_value('3')
        element = driver.find_element(By.XPATH, '//*[@id="main-container"]/div[2]/div/div[2]/div[1]/h1/button')
        element.click()
        time.sleep(5)
    except Exception as e:
        print(f"Error:{e}")
    logger.info('Event Rec On with Video Stream3')

# 이벤트 녹화 사용설정-비디오 스트림4
def recon4():
    try:
        # element = driver.find_element(By.XPATH, '//*[@id="menu_button"]')
        # element.click()
        # time.sleep(5)
        # element = driver.find_element(By.XPATH, '//*[@id="sidebar-shortcuts-large"]/a[3]')
        # element.click()
        # time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/a')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/ul/li[2]/a')
        element.click()
        time.sleep(5)
        select = Select(driver.find_element(By.CSS_SELECTOR,'#real_record_event_video'))
        select.select_by_value('4')
        element = driver.find_element(By.XPATH, '//*[@id="main-container"]/div[2]/div/div[2]/div[1]/h1/button')
        element.click()
        time.sleep(5)
    except Exception as e:
        print(f"Error:{e}")
    logger.info('Event Rec On with Video Stream4')

# 이벤트 녹화 설정(타이머)
def rectimer():
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/a'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/ul/li[1]/a'))).click()
    time.sleep(5)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[1]/ul/li[9]/a'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tab-timer"]/div[1]/label/div'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[9]/div[3]/select'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[9]/div[3]/select/option[1]'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[9]/div[4]/select'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[9]/div[4]/select/option[2]'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button'))).click()
    logger.info('Timer On(1m)')
    time.sleep(5)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/ul/li[3]/a'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[1]/div/div[2]/input[1]'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div[2]/form/div[2]/div[2]/div/div[1]/select'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div[2]/form/div[2]/div[2]/div/div[1]/select/option[46]'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="modal_rule"]/div/div/div[2]/form/div[3]/div[2]/div[8]/div/label/div'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div[3]/button[1]'))).click()
    logger.info('Create Event Rule with Timer')

# 이벤트 영상 추출하기
def videosave(videoname):
    driver.find_element(By.XPATH, '//*[@id="sidebar-shortcuts-large"]/a[2]').click()
    rows = driver.find_elements(By.XPATH, '//tr[@role="row" and @id]')
    nowid = max(int(row.get_attribute('id')) for row in rows)
    elements = driver.find_elements(By.XPATH, f'//tr[@role="row" and @id="{nowid}"]/td[4]')
    if elements:
        elements[0].click()
    else:
        print("WHAT THE FUCK")
    video_element = driver.find_element(By.TAG_NAME, 'video')
    video_url = video_element.get_attribute('src')
    response = requests.get(video_url, timeout=60)
    directory = os.getcwd()
    vid_result = os.path.join(directory, "video_result")
    os.makedirs(vid_result, exist_ok=True)
    file_name = f"{videoname}.mp4"
    file_path = os.path.join(vid_result, file_name)
    file = open(file_path, 'wb') 
    file.write(response.content)
    logger.info("File saved as %s", file_name)
    return vid_result

# 엑셀 저장
def excelsave(vid_result, sheet_name="metadata", excel_path=".\\video_result\\meta.xlsx"):
    data = info(vid_result)
    # 파일이 없으면 새로 만들기
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    else:
        wb = load_workbook(excel_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        # 기존 데이터 지우기 (옵션)
        ws.delete_rows(1, ws.max_row)
    # 데이터 삽입
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(excel_path)
    wb.close()


# 이벤트 영상 조건 확인하기
def info(vid_result):
    metadata_list = []
    header = ["파일명", "코덱 프로파일", "영상 길이", "fps", "width", "height", "비트레이트 모드", "비트레이트"]
    metadata_list.append(header)

    for filename in os.listdir(vid_result):
        if filename.endswith('.mp4'):
            file_path = os.path.join(vid_result, filename)
            media_info = MediaInfo.parse(file_path)
            row = [filename]
            found = False
            for track in media_info.tracks:
                if track.track_type == 'Video':
                    row.extend([
                    row.append(track.format_profile),
                    row.append(track.duration),
                    row.append(track.frame_rate),
                    row.append(track.width),
                    row.append(track.height),
                    row.append(track.bit_rate_mode),
                    row.append(track.bit_rate),
                    ])
                    found = True
                    break
            if not found:
                row.extend([''] * (len(header) - 1))
            metadata_list.append(row)
    return metadata_list

#cama-1
def cama1():
    videoname1 = 'cama1'
    # 압축 방식 H.264 High
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_codec_list'))
    select.select_by_value("H264HP")
    time.sleep(3)
    # 해상도 2160
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_resolution_list'))
    select.select_by_value('3840x2160')
    time.sleep(3)
    # 프레임레이트 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_fps_list'))
    select.select_by_value('30')
    time.sleep(3)
    # GOP크기 60
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_gov_list'))
    select.select_by_value('60')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bit_control_list'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 6000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bps_list'))
    select.select_by_value('6000')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-1 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname1)
    return videoname1

#cama-2
def cama2():
    videoname2 = 'cama2'
    # 압축 방식 H.264 Main
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 1728
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_resolution_list'))
    select.select_by_value('3072x1728')
    time.sleep(3)
    # 프레임레이트 15
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_fps_list'))
    select.select_by_value('15')
    time.sleep(3)
    # GOP크기 10
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_gov_list'))
    select.select_by_value('10')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bit_control_list'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 12000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bps_list'))
    select.select_by_value('12000')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-2 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname2)
    return videoname2

# cama-3 압축방식 Smart
def cama3():
    videoname3 = 'cama3'
    # 압축 방식 H.264 Main
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 1440
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_resolution_list'))
    select.select_by_value('2560x1440')
    time.sleep(3)
    # 프레임레이트 1
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_fps_list'))
    select.select_by_value('1')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bit_control_list'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 1000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bps_list'))
    select.select_by_value('1000')
    time.sleep(3)
    # 비트레이트 품질 10
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc1_bps_quality > select'))
    select.select_by_value('10')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-3 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname3)
    return videoname3

# cama-4 
def cama4():
    videoname4 = 'cama4'
    # 압축 방식 H.264 Smart
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_codec_list'))
    select.select_by_value("H264PL")
    time.sleep(3)
    # 해상도 1440
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_resolution_list'))
    select.select_by_value('2560x1440')
    time.sleep(3)
    # 프레임레이트 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_fps_list'))
    select.select_by_value('30')
    time.sleep(3)
    # GOP크기 10
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_gov_list'))
    select.select_by_value('10')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bit_control_list'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 5000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bps_list'))
    select.select_by_value('5000')
    time.sleep(3)
    # 비트레이트 품질 1
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc1_bps_quality > select'))
    select.select_by_value('1')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-4 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname4)
    return videoname4

# cama-5 ************************비디오스트림2 필수지정!!*******************************
def cama5():
    videoname5 = 'cama5'
    # 압축 방식 H.264 High
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_codec_list'))
    select.select_by_value("H264HP")
    time.sleep(3)
    # 해상도 1080
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_resolution_list'))
    select.select_by_value('1920x1080')
    time.sleep(3)
    # 프레임레이트 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_fps_list'))
    select.select_by_value('30')
    time.sleep(3)
    # GOP크기 60
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_gov_list'))
    select.select_by_value('60')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bit_control_list'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 12000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bps_list'))
    select.select_by_value('12000')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-5 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname5)
    return videoname5

# cama-6
def cama6():
    videoname6 = 'cama6'
# 압축 방식 H.264 Main
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 896
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_resolution_list'))
    select.select_by_value('1600x896')
    time.sleep(3)
    # 프레임레이트 10
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_fps_list'))
    select.select_by_value('10')
    time.sleep(3)
    # GOP크기 1
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_gov_list'))
    select.select_by_value('1')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bit_control_list'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 7500
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bps_list'))
    select.select_by_value('7500')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-6 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname6)
    return videoname6

# cama-7
def cama7():
    videoname7 = 'cama7'
    # 압축 방식 H.264 Smart
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_codec_list'))
    select.select_by_value("H264PL")
    time.sleep(3)
    # 해상도 720
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_resolution_list'))
    select.select_by_value('1280x720')
    time.sleep(3)
    # 프레임레이트 1
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_fps_list'))
    select.select_by_value('1')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bit_control_list'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 100
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bps_list'))
    select.select_by_value('100')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-7 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname7)
    return videoname7

# cama-8
def cama8():
    videoname8 = 'cama8'
    # 압축 방식 H.264 MAIN
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 720
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_resolution_list'))
    select.select_by_value('768x432')
    time.sleep(3)
    # 프레임레이트 1
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_fps_list'))
    select.select_by_value('1')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bit_control_list'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 100
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bps_list'))
    select.select_by_value('100')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-8 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname8)
    return videoname8

# cama-9
def cama9():
    videoname9 = 'cama9'
    # 압축 방식 H.264 MAIN
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 392
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_resolution_list'))
    select.select_by_value('704x392')
    time.sleep(3)
    # 프레임레이트 20
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_fps_list'))
    select.select_by_value('20')
    time.sleep(3)
    # GOP크기 60
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_gov_list'))
    select.select_by_value('60')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bit_control_list'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 1000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bps_list'))
    select.select_by_value('1000')
    time.sleep(3)
    # 비트레이트 품질 1
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc2_bps_quality > select'))
    select.select_by_value('1')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-9 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname9)
    return videoname9

# cama-10
def cama10():
    videoname10 = 'cama10'
    # 압축 방식 H.264 MAIN
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 360
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_resolution_list'))
    select.select_by_value('640x360')
    time.sleep(3)
    # 프레임레이트 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_fps_list'))
    select.select_by_value('30')
    time.sleep(3)
    # GOP크기 60
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_gov_list'))
    select.select_by_value('60')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bit_control_list'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 3000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bps_list'))
    select.select_by_value('3000')
    time.sleep(3)
    # 비트레이트 품질 10
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc2_bps_quality > select'))
    select.select_by_value('10')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-10 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname10)
    return videoname10

# cama-11 !!!!!비디오 스트림 3!!!!!
def cama11():
    videoname11 = 'cama11'
    # 압축 방식 H.264 High
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_codec_list'))
    select.select_by_value("H264HP")
    time.sleep(3)
    # 해상도 432
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_resolution_list'))
    select.select_by_value('768x432')
    time.sleep(3)
    # 프레임레이트 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_fps_list'))
    select.select_by_value('30')
    time.sleep(3)
    # GOP크기 60
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_gov_list'))
    select.select_by_value('60')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_bit_control'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 8000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_bps_list'))
    select.select_by_value('8000')
    time.sleep(3)
    # 비트레이트 품질 10
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc3_bps_quality > select'))
    select.select_by_value('10')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-11 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname11)
    return videoname11

# cama-12
def cama12():
    videoname12 = 'cama12'
    # 압축 방식 H.264 Main
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 392
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_resolution_list'))
    select.select_by_value('704x392')
    time.sleep(3)
    # 프레임레이트 10
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_fps_list'))
    select.select_by_value('10')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_bit_control'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 8000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_bps_list'))
    select.select_by_value('1000')
    time.sleep(3)
    # 비트레이트 품질 10
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc3_bps_quality > select'))
    select.select_by_value('1')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-12 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname12)
    return videoname12

# cama-13
def cama13():
    videoname13 = 'cama13'
    # 압축 방식 H.264 Smart
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_codec_list'))
    select.select_by_value("H264PL")
    time.sleep(3)
    # 해상도 360
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_resolution_list'))
    select.select_by_value('640x360')
    time.sleep(3)
    # 프레임레이트 15
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_fps_list'))
    select.select_by_value('15')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_bit_control'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 800
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_bps_list'))
    select.select_by_value('800')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-13 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname13)
    return videoname13

# cama-14  !!!!!!!비디오스트림4!!!!!!!!!!!
def cama14():
    videoname14 = 'cama14'
    # 압축 방식 H.264 High
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_codec_list'))
    select.select_by_value("H264HP")
    time.sleep(3)
    # 해상도 720
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_resolution_list'))
    select.select_by_value('1280x720')
    time.sleep(3)
    # 프레임레이트 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_fps_list'))
    select.select_by_value('30')
    time.sleep(3)
    # GOP크기 60
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_gov_list'))
    select.select_by_value('60')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bit_control'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 12000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bps_list'))
    select.select_by_value('12000')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-14 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname14)
    return videoname14

def cama15():
    videoname15 = 'cama15'
    # 압축 방식 H.264 High
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 432
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_resolution_list'))
    select.select_by_value('768x432')
    time.sleep(3)
    # 프레임레이트 10
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_fps_list'))
    select.select_by_value('10')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bit_control'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 1000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bps_list'))
    select.select_by_value('1000')
    # 비트레이트 품질 1
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc4_bps_quality > select'))
    select.select_by_value('1')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-15 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname15)
    return videoname15

def cama16():
    videoname16 = 'cama16'
    # 압축 방식 H.264 Main
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 392
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_resolution_list'))
    select.select_by_value('704x392')
    time.sleep(3)
    # 프레임레이트 1
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_fps_list'))
    select.select_by_value('1')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bit_control'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 5000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bps_list'))
    select.select_by_value('5000')
    # 비트레이트 품질 10
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc4_bps_quality > select'))
    select.select_by_value('10')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-16 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname16)
    return videoname16

def cama17():
    videoname17 = 'cama17'
    # 압축 방식 H.264 Main
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 360
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_resolution_list'))
    select.select_by_value('640x360')
    time.sleep(3)
    # 프레임레이트 10
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_fps_list'))
    select.select_by_value('10')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bit_control'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 1000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bps_list'))
    select.select_by_value('1000')
    # 비트레이트 품질 1
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc4_bps_quality > select'))
    select.select_by_value('1')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-17 TC Precondition Set Complete')
    time.sleep(150)
    videosave(videoname17)
    return videoname17

# # 테스트 진행 전 무조건 저장소 초기화, 시스템 초기화 후 비밀번호 초기 설정까지 완료하세요
def main_rec(ip, username, password):
    driver = setup_driver()
    loginop(ip, username, password)
    time.sleep(5)
    recon(driver)
    time.sleep(5)
    reckeep(driver)
    time.sleep(5)
    rectimer(driver)
    time.sleep(5)
    op_stream(driver)
    time.sleep(5)
    cama1(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama2(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama3(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama4(driver)
    time.sleep(5)
    recon2(driver)
    time.sleep(5)
    op_stream(driver)
    time.sleep(5)
    cama5(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)    
    cama6(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama7(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama8(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama9(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama10(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    recon3(driver)
    time.sleep(5)
    op_stream(driver)
    time.sleep(5)
    cama11(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama12(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama13(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    recon4(driver)
    time.sleep(5)
    op_stream(driver)
    time.sleep(5)
    cama14(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama15(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama16(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama17(driver)
    time.sleep(5)

if __name__ == "__main__":
    main_rec()